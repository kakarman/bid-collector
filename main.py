# -*- coding: utf-8 -*-
"""
=============================================================================
 나라장터(조달청) 입찰정보 자동 수집 & 이메일 발송 프로그램
=============================================================================

 [ 이 프로그램이 하는 일 ]
   1. 조달청 나라장터 Open API에 접속해서
   2. 정해진 기간(기본: 전날 오전 10시 ~ 오늘 오전 10시)에 새로 올라온
      [사업계획(발주계획)] / [사전규격] / [본공고] 를 모두 가져오고
   3. 내가 정한 검색어(KEYWORDS)가 들어간 건만 골라내서
   4. 엑셀 파일(.xlsx) 하나로 예쁘게 정리한 뒤
   5. 내 이메일로 첨부해서 보내줍니다.

 [ 코딩을 몰라도 괜찮습니다 ]
   아래 "★★★ 사용자 설정 영역 ★★★" 안의 값만 바꾸면 됩니다.
   그 아래쪽(=== 여기부터는 프로그램 본체 ===)은 건드리지 않아도 됩니다.
=============================================================================
"""

import os                    # 환경변수(깃허브 Secrets)를 읽기 위한 도구
import re                    # 글자 정리(공백 제거 등)를 위한 도구
import ssl                   # 이메일을 안전하게 보내기 위한 도구
import sys                   # 프로그램 종료 처리를 위한 도구
import time                  # 잠깐 쉬어가기(API 과부하 방지)를 위한 도구
import smtplib               # 이메일 발송(SMTP) 도구
import traceback             # 오류가 났을 때 원인을 자세히 보여주는 도구
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage

import requests              # 인터넷(API)에서 데이터를 받아오는 도구
import pandas as pd          # 표(엑셀) 데이터를 다루는 도구


# =============================================================================
# ★★★ 사용자 설정 영역 ★★★  (여기만 편하게 수정하세요)
# =============================================================================

# ---------------------------------------------------------------------------
# 1) 검색어(키워드) 목록  ─ 두 그룹으로 나눠서 관리합니다
#
#    ▸ CORE_KEYWORDS  : 구체적인 장비·제품 이름 (적중률 높음)  → 결과에 ★핵심 표시
#    ▸ WIDE_KEYWORDS  : 분야를 넓게 훑는 포괄적인 단어         → 결과에 참고 표시
#
#    두 그룹 모두 수집은 하되, 엑셀 맨 앞 '중요도' 열로 구분해 드립니다.
#    → 바쁘실 땐 ★핵심만 보시고, 시간 여유 있을 때 참고까지 훑으시면 됩니다.
#
#    [수정 방법]
#      - 추가: 줄을 하나 늘리고  '단어',  형식으로 적으면 됩니다.
#      - 삭제: 그 줄 맨 앞에 # 을 붙이거나 줄을 통째로 지우세요.
#      - 띄어쓰기는 자동 무시됩니다. ('실험 기기' = '실험기기')
# ---------------------------------------------------------------------------

# ★ 핵심 키워드 : 구체적인 장비명 (이게 걸리면 바로 영업 대상)
CORE_KEYWORDS = [
    # 배양·항온 계열
    '인큐베이터',
    'CO2배양기',
    '배양기',
    '배양',
    '항온항습기',
    '항온항습',
    '항온',
    '항습',
    '수조',
    '챔버',

    # 멸균·건조·가열 계열
    '고압멸균',
    '고압증기',
    '오토클레이브',
    '클레이브',
    '멸균기',
    '멸균',
    '건조기',
    '오븐',

    # 냉장·냉동 계열
    '초저온',
    '디프프리저',
    '냉동고',
    '냉장고',
    '제빙기',
    '제빙',

    # 전처리·분석 계열
    '원심분리기',
    '진탕기',
    '쉐이커',
    '교반기',
    '현미경',
    '분광광도계',

    # 클린·작업환경 계열
    '클린벤치',
    '클린룸',
    '무균작업대',
    '안전캐비닛',
    '흄후드',
    '작업대',
    '실험대',

    # 묶음 표현
    '실험기기',
    '실험장비',
    '연구장비',
    '연구기자재',
    '시험장비',
    '분석장비',
    '분석기기',
    '계측장비',
    '이화학',
]

# ○ 넓은 키워드 : 분야를 폭넓게 훑는 단어 (놓치지 않기 위한 그물)
WIDE_KEYWORDS = [
    '실험실',
    '실험',
    '실습',
    '시험',
    '연구',
    '과학',
    '기자재',
    '장비',
    '의료',
    '의무',
    '제약',
    '진단',
    '시약',
    '바이오',
    '생명공학',
    '유전자',
    '세포',
    '생물',
    '미생물',
    '동물',
    '식물',
    '곤충',
    '해양',
    '바다',
    '클린',
    '건조',
]

# 실제 검색에 쓰이는 전체 목록 (건드리지 마세요)
KEYWORDS = CORE_KEYWORDS + WIDE_KEYWORDS

# ---------------------------------------------------------------------------
# 2) 제외 키워드 (선택)
#    - 위 키워드에 걸렸더라도, 아래 단어가 들어있으면 결과에서 빼버립니다.
#    - 예: 청소용역, 경비용역처럼 관련 없는 공고가 자꾸 걸릴 때 사용하세요.
#    - 필요 없으면 아래처럼 비워두면 됩니다:  EXCLUDE_KEYWORDS = []
# ---------------------------------------------------------------------------
EXCLUDE_KEYWORDS = [
    # '청소',
    # '경비용역',
    # '급식',
]

# ---------------------------------------------------------------------------
# 3) 수집할 단계 (True = 수집함 / False = 수집 안 함)
#    - 필요 없는 단계는 False 로 바꾸면 그만큼 빨라집니다.
# ---------------------------------------------------------------------------
COLLECT_STAGES = {
    # ⚠️ 발주계획(사업계획)은 조달청이 API 주소를 개편해서 현재 연결되지 않습니다.
    #    켜두면 매번 접속 실패로 시간만 오래 걸려서 기본값을 False 로 두었습니다.
    #    나중에 정확한 주소를 확인하면 True 로 바꾸시면 됩니다.
    '사업계획': False,  # 발주계획(연간/분기 발주 예정 사업)
    '사전규격': True,   # 규격 사전공개(본공고 나오기 전 단계) ★영업 골든타임
    '본공고': True,     # 실제 입찰공고
}

# ---------------------------------------------------------------------------
# 4) 수집할 업무 구분
#    - 보통 장비 영업은 '물품'과 '용역'이면 충분합니다.
#    - 공사까지 보고 싶으면 '공사' 줄의 # 을 지우세요.
# ---------------------------------------------------------------------------
BUSINESS_TYPES = [
    '물품',
    '용역',
    # '공사',
]

# ---------------------------------------------------------------------------
# 5) 수집 기준 시각 (매일 몇 시를 기준으로 "하루"를 자를지)
#    - 10 이면 "어제 10:00 ~ 오늘 10:00" 입니다.
#    - 깃허브 액션 실행 시간도 같이 바꿔야 합니다(.yml 파일의 cron 참고).
# ---------------------------------------------------------------------------
DAILY_CUTOFF_HOUR = 10

# ---------------------------------------------------------------------------
# 6) 메일 제목 앞에 붙일 말머리
# ---------------------------------------------------------------------------
MAIL_SUBJECT_PREFIX = '[나라장터 입찰정보]'

# ---------------------------------------------------------------------------
# 7) 사전규격 공고링크 설정
#
#    2024년 '차세대 나라장터' 개통 이후 본공고와 사전규격의 상세화면 주소 체계가
#    완전히 달라졌습니다. 본공고는 API가 상세주소를 그대로 주지만(bidNtceDtlUrl),
#    사전규격은 주소를 주지 않는 경우가 있어 아래 규칙으로 직접 만들어 넣습니다.
#
#    [우선순위]
#      1순위: API가 주소를 주면 그대로 사용
#      2순위: 아래 PRESPEC_LINK_TEMPLATE 이 채워져 있으면 그것으로 주소를 생성
#      3순위: 나라장터 사전규격 검색화면 링크 (등록번호로 바로 조회 가능)
#
#    ※ 나중에 사전규격 상세주소 형식을 확인하면, 아래 한 줄만 바꾸면 됩니다.
#      예) PRESPEC_LINK_TEMPLATE = 'https://www.g2b.go.kr/link/○○○/single/?번호={no}'
#          ({no} 자리에 등록번호가 자동으로 들어갑니다)
# ---------------------------------------------------------------------------
PRESPEC_LINK_TEMPLATE = ''

# 사전규격 검색 화면 (등록번호를 붙여넣으면 바로 조회됩니다)
PRESPEC_SEARCH_URL = 'https://www.g2b.go.kr:8341/bs/beffatStndrdUrlSearchList.do?gCode=B553766&cssStyle=3&taskClCd=0'

# ---------------------------------------------------------------------------
# 8) 점검 모드
#    True 로 바꾸면 API가 실제로 어떤 항목(필드)들을 주는지 로그에 찍어줍니다.
#    링크나 예산이 비어 있을 때 원인을 찾는 용도이며, 평소엔 False 로 두세요.
# ---------------------------------------------------------------------------
DEBUG_SHOW_FIELDS = False

# =============================================================================
# ★★★ 사용자 설정 영역 끝 ★★★
#     아래부터는 건드리지 않으셔도 됩니다.
# =============================================================================


# -----------------------------------------------------------------------------
# 기본 상수들
# -----------------------------------------------------------------------------
KST = timezone(timedelta(hours=9))          # 한국 표준시(UTC+9)
API_HOST = 'apis.data.go.kr/1230000'        # 조달청 Open API 공통 주소
API_SCHEMES = ['https', 'http']             # https 우선, 안 되면 http로 재시도
PAGE_SIZE = 500        # 한 번 요청할 때 가져올 건수 (너무 크면 실패할 수 있어 500)
MAX_PAGES = 30         # 한 종류당 최대 페이지 수 (안전장치: 무한루프 방지)
SLEEP_BETWEEN_CALLS = 0.3  # API 호출 사이 쉬는 시간(초) - 과부하 방지

# 조달청 서버가 가끔 응답을 안 주기 때문에(우리 잘못이 아님) 자동으로 다시 시도합니다.
CONNECT_TIMEOUT = 10   # 서버와 연결되기까지 기다리는 시간(초)
READ_TIMEOUT = 60      # 연결된 뒤 데이터를 다 받을 때까지 기다리는 시간(초)
MAX_RETRY = 3          # 통신 실패 시 다시 시도할 횟수
RETRY_WAIT = 3         # 재시도 전 기다리는 시간(초). 실패할수록 2배씩 늘어납니다.

SERVICE_KEY = ''       # 공공데이터포털 인증키 (실행할 때 Secrets 에서 채워집니다)

# 단계별 API 주소 목록
# - 앞의 것부터 시도하고, 실패하면 다음 것으로 자동 전환합니다(자동 백업 경로).
API_ENDPOINTS = {
    '본공고': {
        '물품': ['ad/BidPublicInfoService/getBidPblancListInfoThng',
                 'ad/BidPublicInfoService/getBidPblancListInfoThngPPSSrch'],
        '용역': ['ad/BidPublicInfoService/getBidPblancListInfoServc',
                 'ad/BidPublicInfoService/getBidPblancListInfoServcPPSSrch'],
        '공사': ['ad/BidPublicInfoService/getBidPblancListInfoCnstwk',
                 'ad/BidPublicInfoService/getBidPblancListInfoCnstwkPPSSrch'],
    },
    '사전규격': {
        '물품': ['ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoThng',
                 'ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoThngPPSSrch'],
        '용역': ['ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServc',
                 'ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServcPPSSrch'],
        '공사': ['ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoCnstwk',
                 'ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoCnstwkPPSSrch'],
    },
    # 발주계획(사업계획)은 조달청에서 서비스 주소가 바뀐 적이 있어
    # 가능성 있는 주소를 여러 개 넣어두고 되는 것을 자동으로 찾아 씁니다.
    '사업계획': {
        '물품': ['at/OrderPlanSttusService/getOrderPlanSttusListThng',
                 'ao/OrderPlanSttusService/getOrderPlanSttusListThng',
                 'at/PubPrcrmntOrderPlanService/getOrderPlanSttusListThng'],
        '용역': ['at/OrderPlanSttusService/getOrderPlanSttusListServc',
                 'ao/OrderPlanSttusService/getOrderPlanSttusListServc',
                 'at/PubPrcrmntOrderPlanService/getOrderPlanSttusListServc'],
        '공사': ['at/OrderPlanSttusService/getOrderPlanSttusListCnstwk',
                 'ao/OrderPlanSttusService/getOrderPlanSttusListCnstwk',
                 'at/PubPrcrmntOrderPlanService/getOrderPlanSttusListCnstwk'],
    },
}

# 엑셀에 넣을 열(컬럼) 순서
COLUMNS = [
    '중요도', '구분', '업무', '공고명', '수요기관', '공고기관',
    '등록/공고일시', '사업기한(마감일)', '배정예산(원)',
    '사업내용요약', '매칭키워드', '공고번호', '공고링크',
]

# 중요도 표시 문구
LEVEL_CORE = '★핵심'
LEVEL_WIDE = '참고'

# 발주계획처럼 주소를 못 찾은 서비스를 기억해 두는 곳 (같은 오류 반복 방지)
UNAVAILABLE_STAGES = set()
DEAD_PATHS = set()          # 400/404 가 난 주소 (다시 시도하지 않음)
NETWORK_FAILS = {}          # 주소별 통신 실패 횟수
WARNINGS = []               # 실행 중 생긴 문제 (메일 본문에 함께 알려드립니다)
SESSION = None              # 접속을 재사용해 속도·안정성을 높이는 통신 객체


def add_warning(message):
    """메일로 알려드릴 경고를 기록합니다. (같은 내용은 한 번만)"""
    if message not in WARNINGS:
        WARNINGS.append(message)


def get_session():
    """
    통신 담당 객체를 만듭니다.
    - 접속을 재사용해서 매번 새로 연결하는 낭비를 줄이고
    - 서버가 일시적으로 먹통일 때(500/502/503/504, 연결 끊김) 자동으로 재시도합니다.
    """
    global SESSION
    if SESSION is not None:
        return SESSION

    from requests.adapters import HTTPAdapter
    try:
        from urllib3.util.retry import Retry
    except ImportError:                      # 아주 오래된 환경 대비
        from requests.packages.urllib3.util.retry import Retry

    retry_rule = Retry(
        total=MAX_RETRY,
        connect=MAX_RETRY,
        read=2,
        backoff_factor=1.5,                  # 1.5초 → 3초 → 6초 간격으로 재시도
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(['GET']),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_rule, pool_connections=4, pool_maxsize=8)

    SESSION = requests.Session()
    SESSION.mount('https://', adapter)
    SESSION.mount('http://', adapter)
    SESSION.headers.update({'User-Agent': 'bid-collector/1.0'})
    return SESSION


# -----------------------------------------------------------------------------
# [도우미 함수] 로그(진행상황) 출력
# -----------------------------------------------------------------------------
def log(message):
    """실행 중 무슨 일이 일어나는지 화면에 찍어줍니다. (깃허브 액션 로그에서 확인 가능)"""
    now = datetime.now(KST).strftime('%H:%M:%S')
    print(f'[{now}] {message}', flush=True)


# -----------------------------------------------------------------------------
# [도우미 함수] 환경변수(깃허브 Secrets) 읽기
# -----------------------------------------------------------------------------
def get_env(name, required=True, default=''):
    """
    깃허브 Secrets 에 등록한 값을 읽어옵니다.
    required=True 인데 값이 없으면 프로그램을 즉시 멈추고 이유를 알려줍니다.
    """
    value = os.environ.get(name, default)
    value = (value or '').strip()
    if required and not value:
        log(f'❌ 필수 설정값 "{name}" 이(가) 비어 있습니다. 깃허브 Secrets 를 확인하세요.')
        sys.exit(1)
    return value


# -----------------------------------------------------------------------------
# [핵심 1] 수집 기간 계산하기
# -----------------------------------------------------------------------------
def calc_period(now=None, lookback_days=0):
    """
    '언제부터 언제까지' 올라온 공고를 가져올지 계산합니다.

    기본 규칙
      - 끝 시각  : 지금(실행 시각)
      - 시작 시각: 어제 오전 10시  (DAILY_CUTOFF_HOUR 값)
      - 월요일이면: 금요일 오전 10시부터 (→ 금·토·일 공고를 모두 포함)

    lookback_days 가 1 이상이면(수동 실행 시 입력) 그 일수만큼 거슬러 올라갑니다.
    """
    if now is None:
        now = datetime.now(KST)

    end_dt = now

    # (1) 수동 실행에서 "최근 N일" 을 지정한 경우 → 그 값을 최우선으로 사용
    if lookback_days and lookback_days > 0:
        start_dt = (now - timedelta(days=lookback_days)).replace(
            hour=DAILY_CUTOFF_HOUR, minute=0, second=0, microsecond=0)
        return start_dt, end_dt

    # (2) 기본 로직: 오늘 기준 며칠 전부터 볼 것인가?
    #     weekday(): 월=0, 화=1, 수=2, 목=3, 금=4, 토=5, 일=6
    weekday = now.weekday()
    if weekday == 0:        # 월요일 → 3일 전(금요일)부터
        days_back = 3
    elif weekday == 6:      # 일요일 → 2일 전(금요일)부터 (주말 수동 실행 대비)
        days_back = 2
    else:                   # 화~토 → 하루 전부터
        days_back = 1

    start_dt = (now - timedelta(days=days_back)).replace(
        hour=DAILY_CUTOFF_HOUR, minute=0, second=0, microsecond=0)

    # 안전장치: 혹시 시작이 끝보다 뒤면(새벽 실행 등) 하루 더 뒤로 밀어줍니다.
    if start_dt >= end_dt:
        start_dt = start_dt - timedelta(days=1)

    return start_dt, end_dt


def fmt_api_datetime(dt):
    """API가 요구하는 날짜 형식(yyyyMMddHHmm)으로 바꿔줍니다. 예) 202608091000"""
    return dt.strftime('%Y%m%d%H%M')


# -----------------------------------------------------------------------------
# [핵심 2] API 호출
# -----------------------------------------------------------------------------
def clean_service_key(raw_key):
    """
    공공데이터포털 인증키는 '일반 인증키(Decoding)' 를 써야 합니다.
    혹시 Encoding 키(%2B, %3D 같은 게 섞인 키)를 넣었다면 자동으로 풀어줍니다.
    """
    from urllib.parse import unquote
    key = raw_key.strip()
    if '%' in key:
        key = unquote(key)
    return key


def call_api(path, params, page_size):
    """
    조달청 API를 1회 호출하고 결과(JSON)를 돌려줍니다.
    실패하면 None 을 돌려주고, 이유를 로그에 남깁니다.
    """
    query = dict(params)
    query['numOfRows'] = page_size
    query['type'] = 'json'

    op_name = path.split('/')[-1]
    response = None
    last_error = None

    # https → http 순서로, 각각 여러 번 시도합니다.
    # (조달청 서버는 접속이 몰리면 응답을 아예 안 주는 경우가 있습니다)
    for scheme in API_SCHEMES:
        url = f'{scheme}://{API_HOST}/{path}'
        for attempt in range(1, MAX_RETRY + 1):
            try:
                response = get_session().get(
                    url, params=query, timeout=(CONNECT_TIMEOUT, READ_TIMEOUT))
                last_error = None
                break
            except Exception as error:
                last_error = error
                response = None
                if attempt < MAX_RETRY:
                    wait = RETRY_WAIT * (2 ** (attempt - 1))   # 3초 → 6초 → 12초
                    log(f'   ⏳ 응답이 없어 {wait}초 뒤 다시 시도합니다 '
                        f'({op_name}, {attempt}/{MAX_RETRY})')
                    time.sleep(wait)
        if response is not None:
            break

    if response is None:
        # 여러 번 시도해도 실패 → 이 주소는 이번 실행에서 더 시도하지 않습니다.
        fails = NETWORK_FAILS.get(path, 0) + 1
        NETWORK_FAILS[path] = fails
        reason = type(last_error).__name__ if last_error else '알 수 없음'
        log(f'   ⚠️ 통신 실패({op_name}) - {reason}. 조달청 서버 응답 없음.')
        if fails >= 2:
            DEAD_PATHS.add(path)
            add_warning(f'{op_name}: 조달청 서버 접속 실패로 이번 회차는 건너뛰었습니다.')
        return None

    if response.status_code != 200:
        # 400/404 는 대부분 "그 주소의 서비스가 없다"는 뜻이므로
        # 한 번만 알리고 다음부터는 이 주소를 아예 시도하지 않습니다.
        if response.status_code in (400, 404):
            if path not in DEAD_PATHS:
                snippet = re.sub(r'<[^>]+>', ' ', response.text[:200]).strip()
                log(f'   ⓘ 미제공 주소 [{response.status_code}] '
                    f'{path.split("/")[-1]} {snippet[:80]}')
            DEAD_PATHS.add(path)
        else:
            log(f'   ⚠️ 서버 응답코드 {response.status_code} ({path.split("/")[-1]})')
        return None

    # 정상이면 JSON, 문제가 있으면 XML 형태의 오류문이 오는 경우가 있습니다.
    try:
        data = response.json()
    except Exception:
        snippet = response.text[:250].replace('\n', ' ')
        log(f'   ⚠️ JSON이 아닌 응답이 왔습니다 ({path}) → {snippet}')
        return None

    # 응답 안의 결과코드 확인 ('00' 이면 정상)
    header = (data.get('response', {}) or {}).get('header', {}) or {}
    result_code = str(header.get('resultCode', '')).strip()
    if result_code and result_code not in ('00', '0'):
        log(f'   ⚠️ API 오류 [{result_code}] {header.get("resultMsg", "")} ({path})')
        return None

    return data


def extract_items(data):
    """API 응답 덩어리에서 실제 목록(items)만 꺼내옵니다."""
    body = (data.get('response', {}) or {}).get('body', {}) or {}
    items = body.get('items', [])
    if isinstance(items, dict):          # {'item': [...]} 형태인 경우
        items = items.get('item', [])
    if isinstance(items, dict):          # 결과가 1건이면 리스트가 아닐 수 있음
        items = [items]
    if not isinstance(items, list):
        items = []
    total = body.get('totalCount', len(items))
    try:
        total = int(total)
    except Exception:
        total = len(items)
    return items, total


def fetch_all(path, start_dt, end_dt):
    """
    한 종류(예: 본공고-물품)의 데이터를 기간 내 전부 가져옵니다.
    페이지를 넘겨가며(1페이지, 2페이지 ...) 끝까지 수집합니다.
    """
    collected = []
    page_size = PAGE_SIZE

    for page_no in range(1, MAX_PAGES + 1):
        params = {
            'serviceKey': SERVICE_KEY,
            'pageNo': page_no,
            'inqryDiv': '1',                        # 1 = 등록/공고 일시 기준 조회
            'inqryBgnDt': fmt_api_datetime(start_dt),
            'inqryEndDt': fmt_api_datetime(end_dt),
        }

        data = call_api(path, params, page_size)

        # 한 번에 500건 요청이 막히는 서버도 있어, 실패 시 100건으로 재시도합니다.
        if data is None and page_size != 100:
            page_size = 100
            data = call_api(path, params, page_size)

        if data is None:
            break

        items, total = extract_items(data)
        collected.extend(items)

        # 더 가져올 게 없으면 종료
        if len(items) < page_size or len(collected) >= total:
            break

        time.sleep(SLEEP_BETWEEN_CALLS)

    return collected


def fetch_with_fallback(stage, biz_type, start_dt, end_dt):
    """
    한 단계·업무구분에 대해 주소 후보를 차례로 시도해서
    데이터가 나오는 주소를 자동으로 찾아 사용합니다.
    """
    # 이미 "이 서비스는 주소 자체가 없다"고 판명된 단계는 조용히 건너뜁니다.
    if stage in UNAVAILABLE_STAGES:
        return []

    all_paths = API_ENDPOINTS.get(stage, {}).get(biz_type, [])
    paths = [p for p in all_paths if p not in DEAD_PATHS]

    items = []
    for path in paths:
        items = fetch_all(path, start_dt, end_dt)
        if items:
            log(f'   ✅ {stage}-{biz_type}: {len(items)}건 수신 ({path.split("/")[-1]})')
            return items

    # 이 단계에 등록된 모든 주소가 "없는 주소"로 판명되면 단계 전체를 접습니다.
    stage_paths = [p for paths_by_biz in API_ENDPOINTS.get(stage, {}).values()
                   for p in paths_by_biz]
    if stage_paths and all(p in DEAD_PATHS for p in stage_paths):
        UNAVAILABLE_STAGES.add(stage)
        log(f'   ⏭  {stage}: 조달청이 제공하지 않는 주소라 이 단계는 건너뜁니다.')
        return []

    log(f'   · {stage}-{biz_type}: 수신 0건')
    return []


# -----------------------------------------------------------------------------
# [핵심 3] 키워드 걸러내기
# -----------------------------------------------------------------------------
def normalize(text):
    """
    비교하기 좋게 글자를 정리합니다.
    괄호·쉼표 같은 기호는 '띄어쓰기'로 바꾸고, 연속된 공백은 하나로 줄입니다.

    ※ 기호를 '삭제'하지 않고 '공백'으로 바꾸는 이유:
       예) '강릉~제진 단선전철' 에서 공백을 지워버리면 '제진단선' 이 되어
           엉뚱하게 '진단' 키워드에 걸립니다. 공백을 남겨두면 이런 오탐이 사라집니다.
    """
    if text is None:
        return ''
    cleaned = re.sub(r'[\-_/()\[\]·,.~:;|]', ' ', str(text)).lower()
    return re.sub(r'\s+', ' ', cleaned).strip()


def strip_spaces(text):
    """띄어쓰기를 모두 없앤 형태 (긴 키워드 비교에만 사용)"""
    return re.sub(r'\s+', '', text)


NORMALIZED_KEYWORDS = [(kw, normalize(kw)) for kw in KEYWORDS]
NORMALIZED_EXCLUDES = [normalize(kw) for kw in EXCLUDE_KEYWORDS if kw.strip()]
CORE_KEYWORD_SET = set(CORE_KEYWORDS)


def decide_level(hits):
    """걸린 키워드 중에 '핵심 장비명'이 하나라도 있으면 ★핵심으로 표시합니다."""
    for keyword in hits:
        if keyword in CORE_KEYWORD_SET:
            return LEVEL_CORE
    return LEVEL_WIDE


def match_keywords(text):
    """
    주어진 글자 안에 내 키워드가 있는지 확인합니다.
    - 걸린 키워드 목록을 돌려주고, 하나도 없으면 빈 목록을 돌려줍니다.
    - 제외 키워드가 걸리면 무조건 빈 목록(=수집 안 함)입니다.
    """
    haystack = normalize(text)
    if not haystack:
        return []
    haystack_tight = strip_spaces(haystack)

    for exclude in NORMALIZED_EXCLUDES:
        if exclude and exclude in haystack:
            return []

    hits = []
    for original, norm in NORMALIZED_KEYWORDS:
        if not norm:
            continue
        if norm in haystack:
            hits.append(original)
            continue
        # 5글자 이상인 긴 키워드는 띄어쓰기가 달라도 찾아줍니다.
        # (예: 키워드 '무균작업대'  ↔  공고명 '무균 작업대')
        # 짧은 단어까지 이렇게 하면 엉뚱한 게 걸려서 길이 제한을 뒀습니다.
        tight = strip_spaces(norm)
        if len(tight) >= 5 and tight in haystack_tight:
            hits.append(original)
    return hits


# -----------------------------------------------------------------------------
# [핵심 4] API 응답을 엑셀용 표 형태로 정리
# -----------------------------------------------------------------------------
def pick(item, candidate_keys, default=''):
    """
    API 응답의 항목 이름이 서비스마다 조금씩 달라서,
    가능한 이름들을 순서대로 찾아보고 값이 있는 것을 씁니다.
    """
    for key in candidate_keys:
        value = item.get(key)
        if value not in (None, '', ' '):
            return str(value).strip()
    return default


def fmt_datetime_text(text):
    """20260810 1000 / 202608101000 같은 값을 '2026-08-10 10:00' 형태로 보기 좋게."""
    digits = re.sub(r'\D', '', str(text or ''))
    if len(digits) >= 12:
        return f'{digits[0:4]}-{digits[4:6]}-{digits[6:8]} {digits[8:10]}:{digits[10:12]}'
    if len(digits) >= 8:
        return f'{digits[0:4]}-{digits[4:6]}-{digits[6:8]}'
    return str(text or '').strip()


def to_amount(text):
    """'1,200,000원' 같은 값을 숫자 1200000 으로 바꿔줍니다. 못 바꾸면 빈칸."""
    digits = re.sub(r'[^\d]', '', str(text or ''))
    if not digits:
        return ''
    try:
        return int(digits)
    except Exception:
        return ''


def build_summary(item, stage):
    """
    '사업내용 요약' 열을 만듭니다.
    나라장터 API는 공고 본문 전체를 주지 않기 때문에,
    API가 제공하는 항목들(품명/규격/계약방법/낙찰방법 등)을 모아 요약으로 씁니다.
    """
    parts = []

    def add(label, keys):
        value = pick(item, keys)
        if value:
            parts.append(f'{label}: {value}')

    add('품명', ['prdctClsfcNoNm', 'prdctIdntNoNm', 'prdctNm', 'bsnsNm'])
    add('규격', ['prdctSpecNm', 'prdctSpec', 'specNm'])
    add('수량', ['prdctQty', 'qty'])
    add('계약방법', ['cntrctCnclsMthdNm', 'cntrctMthdNm', 'cntrctCnclsSttusNm'])
    add('낙찰방법', ['sucsfbidMthdNm', 'sucsfbidLwltRate'])
    add('참가자격', ['bidprcPsblIndstrytyNm', 'prtcptPsblRgnNm', 'indstrytyNm'])
    add('발주예정시기', ['orderPlanDt', 'orderPlanYm', 'cntrctMth', 'ordrPlanMt'])
    add('비고', ['ntceKindNm', 'rgstTyNm', 'bidNtceDtlsNm', 'bsnsDivNm'])

    if stage == '사전규격':
        add('의견등록마감', ['opninRgstClseDt'])

    return ' | '.join(parts)


def build_link(item, stage):
    """
    공고 상세 링크를 만듭니다.

    [1순위] API가 상세주소를 준 경우 → 그대로 사용
    [2순위] 사전규격이면 → 등록번호로 주소를 직접 조립
    [3순위] 그래도 없으면 → 사전규격 검색화면 주소 (절대 빈칸으로 두지 않습니다)
    """
    # ── 1순위: API가 주는 주소 (이름이 서비스마다 달라서 후보를 넉넉히 둡니다)
    url = pick(item, [
        'bidNtceDtlUrl',        # 본공고 상세 URL
        'bfSpecRgstDtlUrl',     # 사전규격 상세 URL
        'specDtlUrl',
        'stdDtlUrl',
        'prestdDtlUrl',
        'dtlUrl',
        'srvceDtlUrl',
        'ntceSpecDocUrl1',      # 규격서 파일 URL
        'specDocUrl',
    ])
    if url.startswith('http'):
        return url

    # ── 2·3순위: 사전규격은 반드시 클릭 가능한 주소를 채워 넣습니다.
    if stage == '사전규격':
        reg_no = pick(item, [
            'bfSpecRgstNo',     # 사전규격등록번호
            'specRgstNo',
            'befatStdrdNo',
            'rgstNo',
        ])
        if reg_no and PRESPEC_LINK_TEMPLATE:
            return PRESPEC_LINK_TEMPLATE.replace('{no}', reg_no)
        return PRESPEC_SEARCH_URL

    return url


def parse_item(item, stage, biz_type):
    """API 응답 한 건(item)을 엑셀 한 줄(dict)로 바꿉니다."""
    title = pick(item, [
        'bidNtceNm',            # 본공고: 공고명
        'prdctClsfcNoNm',       # 사전규격: 품명
        'bsnsNm',               # 발주계획: 사업명
        'ntceNm', 'prdctNm', 'sbmsnRcptNm',
    ])
    summary = build_summary(item, stage)

    # 공고명 + 사업내용요약 을 대상으로 키워드를 찾습니다.
    # (기관명은 일부러 제외합니다. '○○생명공학연구원' 같은 이름 때문에
    #  그 기관의 모든 공고가 딸려오는 것을 막기 위해서입니다.)
    org = pick(item, ['dminsttNm', 'rlDminsttNm', 'demandInsttNm', 'orderInsttNm'])
    hits = match_keywords(f'{title} {summary}')
    if not hits:
        return None     # 키워드에 안 걸리면 버립니다.

    row = {
        '중요도': decide_level(hits),
        '구분': stage,
        '업무': biz_type,
        '공고명': title,
        '수요기관': org,
        '공고기관': pick(item, ['ntceInsttNm', 'orderInsttNm', 'insttNm']),
        '등록/공고일시': fmt_datetime_text(pick(item, [
            'bidNtceDt', 'rgstDt', 'rcptDt', 'bfSpecRgstDt', 'orderPlanRgstDt',
        ])),
        '사업기한(마감일)': fmt_datetime_text(pick(item, [
            'bidClseDt',            # 본공고 입찰마감일시
            'opninRgstClseDt',      # 사전규격 의견등록 마감
            'opengDt',              # 개찰일시
            'bidBeginDt',
            'orderPlanDt', 'ordrPlanMt',
        ])),
        '배정예산(원)': to_amount(pick(item, [
            'asignBdgtAmt',         # 배정예산액
            'presmptPrce',          # 추정가격
            'bdgtAmt', 'sumOfBdgt', 'orderPlanAmt', 'budgetAmount',
        ])),
        '사업내용요약': summary,
        '매칭키워드': ', '.join(hits),
        '공고번호': pick(item, [
            'bidNtceNo', 'bfSpecRgstNo', 'specRgstNo', 'orderPlanNo', 'refNo',
        ]),
        '공고링크': build_link(item, stage),
    }
    return row


# -----------------------------------------------------------------------------
# [핵심 5] 엑셀 파일 만들기
# -----------------------------------------------------------------------------
def make_excel(df, file_path, start_dt, end_dt):
    """정리된 표를 보기 좋은 엑셀 파일로 저장합니다."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='입찰정보', startrow=1)

        workbook = writer.book
        sheet = writer.sheets['입찰정보']

        # --- 맨 윗줄: 수집 기간 안내 ---
        core_count = int((df['중요도'] == LEVEL_CORE).sum()) if '중요도' in df else 0
        sheet.cell(row=1, column=1).value = (
            f'수집기간: {start_dt.strftime("%Y-%m-%d %H:%M")} ~ '
            f'{end_dt.strftime("%Y-%m-%d %H:%M")}  |  총 {len(df)}건 '
            f'(★핵심 {core_count}건 / 참고 {len(df) - core_count}건)  '
            f'※ 중요도 열의 필터를 눌러 ★핵심만 볼 수 있습니다'
        )
        sheet.cell(row=1, column=1).font = Font(bold=True, size=11)

        # --- 머리글(2행) 꾸미기 ---
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # --- 열 너비 지정 ---
        widths = {
            '중요도': 9, '구분': 10, '업무': 8, '공고명': 55, '수요기관': 22, '공고기관': 22,
            '등록/공고일시': 18, '사업기한(마감일)': 18, '배정예산(원)': 16,
            '사업내용요약': 60, '매칭키워드': 22, '공고번호': 18, '공고링크': 35,
        }
        for idx, column_name in enumerate(df.columns, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = widths.get(column_name, 18)

        # --- 본문 셀 정렬 / 금액 서식 / 링크 처리 ---
        columns = list(df.columns)
        amount_col = columns.index('배정예산(원)') + 1 if '배정예산(원)' in columns else None
        link_col = columns.index('공고링크') + 1 if '공고링크' in columns else None
        stage_col = columns.index('구분') + 1 if '구분' in columns else None
        level_col = columns.index('중요도') + 1 if '중요도' in columns else None

        stage_colors = {
            '사업계획': 'FFF2CC',   # 연노랑
            '사전규격': 'DDEBF7',   # 연파랑
            '본공고': 'E2EFDA',     # 연초록
        }

        for row_idx in range(3, len(df) + 3):
            for col_idx in range(1, len(columns) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
                cell.font = Font(size=10)

            if amount_col:
                sheet.cell(row=row_idx, column=amount_col).number_format = '#,##0'
                sheet.cell(row=row_idx, column=amount_col).alignment = Alignment(
                    horizontal='right', vertical='top')

            if stage_col:
                stage_cell = sheet.cell(row=row_idx, column=stage_col)
                color = stage_colors.get(str(stage_cell.value), None)
                if color:
                    stage_cell.fill = PatternFill('solid', fgColor=color)
                stage_cell.alignment = Alignment(horizontal='center', vertical='top')

            if level_col:
                level_cell = sheet.cell(row=row_idx, column=level_col)
                if str(level_cell.value) == LEVEL_CORE:
                    level_cell.fill = PatternFill('solid', fgColor='FFD966')  # 진노랑
                    level_cell.font = Font(size=10, bold=True, color='7F6000')
                else:
                    level_cell.fill = PatternFill('solid', fgColor='F2F2F2')  # 회색
                    level_cell.font = Font(size=10, color='808080')
                level_cell.alignment = Alignment(horizontal='center', vertical='top')

            if link_col:
                link_cell = sheet.cell(row=row_idx, column=link_col)
                url = str(link_cell.value or '').strip()
                if url.startswith('http'):
                    link_cell.hyperlink = url
                    # 사전규격 검색화면으로 보내는 경우엔 문구를 다르게 표시합니다.
                    if url == PRESPEC_SEARCH_URL:
                        link_cell.value = '사전규격 검색(번호로 조회)'
                    else:
                        link_cell.value = '공고 바로가기'
                    link_cell.font = Font(color='0563C1', underline='single', size=10)

        # --- 필터 + 틀 고정 ---
        last_col_letter = get_column_letter(len(columns))
        sheet.auto_filter.ref = f'A2:{last_col_letter}{len(df) + 2}'
        sheet.freeze_panes = 'D3'   # 중요도/구분/업무 열은 스크롤해도 항상 보이게

    return file_path


# -----------------------------------------------------------------------------
# [핵심 6] 이메일 발송
# -----------------------------------------------------------------------------
def warning_html():
    """조달청 서버 문제 등으로 일부를 못 가져왔을 때 메일 맨 위에 알려줍니다."""
    if not WARNINGS:
        return ''
    items = ''.join(f'<li>{w}</li>' for w in WARNINGS)
    return f"""
      <div style="background:#FFF4E5;border-left:4px solid #F0A030;
                  padding:10px 14px;margin:10px 0;">
        <b>⚠️ 참고: 이번 수집 중 일부 문제가 있었습니다</b>
        <ul style="margin:6px 0 0 0;padding-left:18px;">{items}</ul>
        <div style="color:#666;font-size:12px;margin-top:6px;">
          조달청 서버 사정으로 생기는 일시적 현상인 경우가 많습니다.
          누락이 걱정되시면 Actions에서 수동 실행(lookback_days=2)을 한 번 해주세요.
        </div>
      </div>"""


def build_mail_body(df, start_dt, end_dt):
    """메일 본문(HTML)을 만듭니다. 엑셀을 열지 않아도 대충 볼 수 있게 해줍니다."""
    period_text = (f'{start_dt.strftime("%Y-%m-%d %H:%M")} ~ '
                   f'{end_dt.strftime("%Y-%m-%d %H:%M")}')

    if df.empty:
        return f"""
        <div style="font-family:'맑은 고딕',sans-serif;font-size:14px;">
          <p>안녕하세요. 나라장터 입찰정보 자동수집 결과입니다.</p>
          {warning_html()}
          <p><b>수집기간:</b> {period_text}</p>
          <p style="color:#c00;"><b>해당 기간에 키워드와 일치하는 신규 공고가 없습니다.</b></p>
        </div>
        """

    # 단계별 / 중요도별 건수 요약
    counts = df['구분'].value_counts().to_dict()
    summary_line = ' / '.join(f'{k} {v}건' for k, v in counts.items())
    core_df = df[df['중요도'] == LEVEL_CORE]
    core_count = len(core_df)

    # 메일 본문에는 ★핵심 건만 보여줍니다. (핵심이 없으면 전체에서 상위 30건)
    preview_df = core_df if core_count else df

    rows_html = ''
    for _, row in preview_df.head(30).iterrows():
        budget = row['배정예산(원)']
        budget_text = f'{budget:,}' if isinstance(budget, int) else '-'
        link = str(row['공고링크'] or '')
        link_html = f'<a href="{link}">보기</a>' if link.startswith('http') else '-'
        rows_html += f"""
          <tr>
            <td style="border:1px solid #ddd;padding:6px;white-space:nowrap;">{row['구분']}</td>
            <td style="border:1px solid #ddd;padding:6px;">{row['공고명']}</td>
            <td style="border:1px solid #ddd;padding:6px;white-space:nowrap;">{row['사업기한(마감일)'] or '-'}</td>
            <td style="border:1px solid #ddd;padding:6px;text-align:right;white-space:nowrap;">{budget_text}</td>
            <td style="border:1px solid #ddd;padding:6px;text-align:center;">{link_html}</td>
          </tr>"""

    more_note = ('<p style="color:#666;">※ 위 표는 일부만 보여드린 것입니다. '
                 '전체 내용은 첨부된 엑셀 파일을 확인해 주세요. '
                 '엑셀의 <b>중요도</b> 열 필터에서 <b>★핵심</b>만 골라 보실 수 있습니다.</p>')

    return f"""
    <div style="font-family:'맑은 고딕',sans-serif;font-size:14px;">
      <p>안녕하세요. 나라장터 입찰정보 자동수집 결과입니다.</p>
      {warning_html()}
      <p><b>수집기간:</b> {period_text}<br>
         <b>총 건수:</b> {len(df)}건
         (<b style="color:#c55a11;">★핵심 {core_count}건</b> / 참고 {len(df) - core_count}건)<br>
         <b>단계별:</b> {summary_line}</p>
      <p style="margin-bottom:4px;"><b>
        {'★핵심 공고 미리보기' if core_count else '수집 결과 미리보기'}
      </b></p>
      <table style="border-collapse:collapse;font-size:13px;">
        <thead>
          <tr style="background:#1F4E79;color:#fff;">
            <th style="border:1px solid #ddd;padding:6px;">구분</th>
            <th style="border:1px solid #ddd;padding:6px;">공고명</th>
            <th style="border:1px solid #ddd;padding:6px;">사업기한</th>
            <th style="border:1px solid #ddd;padding:6px;">배정예산(원)</th>
            <th style="border:1px solid #ddd;padding:6px;">링크</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
      {more_note}
      <p style="color:#888;font-size:12px;margin-top:20px;">
        본 메일은 GitHub Actions 자동화 스크립트가 발송했습니다.
      </p>
    </div>
    """


def send_email(subject, html_body, attachment_path=None):
    """Gmail SMTP 를 통해 메일을 보냅니다."""
    sender = get_env('GMAIL_USER')                 # 보내는 사람 (내 Gmail 주소)
    password = get_env('GMAIL_APP_PASSWORD')       # Gmail '앱 비밀번호' 16자리
    receivers = get_env('MAIL_TO')                 # 받는 사람 (쉼표로 여러 명 가능)

    receiver_list = [addr.strip() for addr in receivers.replace(';', ',').split(',')
                     if addr.strip()]

    message = EmailMessage()
    message['Subject'] = subject
    message['From'] = sender
    message['To'] = ', '.join(receiver_list)
    message.set_content('HTML 메일입니다. HTML을 지원하는 메일앱에서 확인해 주세요.')
    message.add_alternative(html_body, subtype='html')

    # 엑셀 파일 첨부
    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, 'rb') as file:
            message.add_attachment(
                file.read(),
                maintype='application',
                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=os.path.basename(attachment_path),
            )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as server:
        # Gmail 앱 비밀번호는 공백 없이 16자리입니다. 혹시 공백이 있으면 제거.
        server.login(sender, password.replace(' ', ''))
        server.send_message(message)

    log(f'📧 메일 발송 완료 → {", ".join(receiver_list)}')


# -----------------------------------------------------------------------------
# [메인] 실제 실행 순서
# -----------------------------------------------------------------------------
def main():
    global SERVICE_KEY

    log('=' * 60)
    log('나라장터 입찰정보 자동 수집을 시작합니다.')
    log('=' * 60)

    # (0) 공공데이터포털 인증키 읽기
    SERVICE_KEY = clean_service_key(get_env('SERVICE_KEY'))

    # (1) 수집 기간 계산
    lookback_days = 0
    try:
        lookback_days = int(os.environ.get('LOOKBACK_DAYS', '0') or 0)
    except ValueError:
        lookback_days = 0

    start_dt, end_dt = calc_period(lookback_days=lookback_days)
    log(f'🗓  수집기간: {start_dt.strftime("%Y-%m-%d(%a) %H:%M")} '
        f'~ {end_dt.strftime("%Y-%m-%d(%a) %H:%M")}')
    if end_dt.weekday() == 0 and not lookback_days:
        log('   (월요일이므로 금·토·일 공고까지 함께 수집합니다)')

    # (2) 단계별로 데이터 수집
    rows = []
    raw_total = 0

    for stage, enabled in COLLECT_STAGES.items():
        if not enabled:
            log(f'⏭  {stage}: 설정에서 꺼져 있어 건너뜁니다.')
            continue

        log(f'🔍 [{stage}] 수집 중...')
        for biz_type in BUSINESS_TYPES:
            try:
                items = fetch_with_fallback(stage, biz_type, start_dt, end_dt)
            except Exception as error:
                log(f'   ⚠️ {stage}-{biz_type} 수집 중 오류: {error}')
                items = []

            raw_total += len(items)

            # 점검 모드: API가 실제로 어떤 항목을 주는지 첫 1건만 펼쳐 보여줍니다.
            if DEBUG_SHOW_FIELDS and items and isinstance(items[0], dict):
                log(f'   🔧 [{stage}-{biz_type}] API가 주는 항목 목록:')
                for key in sorted(items[0].keys()):
                    sample = str(items[0].get(key, ''))[:60]
                    log(f'        {key} = {sample}')

            for item in items:
                if not isinstance(item, dict):
                    continue
                row = parse_item(item, stage, biz_type)
                if row:
                    rows.append(row)

    log(f'📊 전체 수신 {raw_total}건 → 키워드 일치 {len(rows)}건')
    if WARNINGS:
        log('⚠️ 이번 실행 중 아래 문제가 있었습니다 (메일에도 함께 안내됩니다)')
        for warning in WARNINGS:
            log(f'    - {warning}')

    # (3) 표로 정리 + 중복 제거 + 정렬
    if rows:
        df = pd.DataFrame(rows, columns=COLUMNS)
        df = df.drop_duplicates(subset=['구분', '공고번호', '공고명'], keep='first')
        # ★핵심을 맨 위로 올리고,
        # 그 안에서 진행 단계(사업계획 → 사전규격 → 본공고),
        # 다시 그 안에서 최신 공고 순으로 정렬합니다.
        stage_order = {'사업계획': 0, '사전규격': 1, '본공고': 2}
        level_order = {LEVEL_CORE: 0, LEVEL_WIDE: 1}
        df['_중요'] = df['중요도'].map(level_order).fillna(9)
        df['_순서'] = df['구분'].map(stage_order).fillna(9)
        df = df.sort_values(
            by=['_중요', '_순서', '등록/공고일시'], ascending=[True, True, False]
        ).drop(columns=['_중요', '_순서']).reset_index(drop=True)
    else:
        df = pd.DataFrame(columns=COLUMNS)

    # (4) 엑셀 만들기
    file_name = f'입찰정보_{end_dt.strftime("%Y%m%d_%H%M")}.xlsx'
    attachment = None
    if not df.empty:
        make_excel(df, file_name, start_dt, end_dt)
        attachment = file_name
        log(f'📁 엑셀 생성 완료: {file_name}')
    else:
        log('📁 수집 결과가 없어 엑셀은 만들지 않습니다.')

    # (5) 메일 발송
    core_count = int((df['중요도'] == LEVEL_CORE).sum()) if not df.empty else 0
    subject = (f'{MAIL_SUBJECT_PREFIX} {end_dt.strftime("%m월 %d일")} '
               f'★핵심 {core_count}건 / 전체 {len(df)}건')
    body = build_mail_body(df, start_dt, end_dt)
    send_email(subject, body, attachment)

    log('✅ 모든 작업이 정상적으로 끝났습니다.')


# 이 파일을 직접 실행했을 때만 main() 이 돌아갑니다.
if __name__ == '__main__':
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        # 예상 못 한 오류가 나면 원인을 로그에 남기고, 오류 알림 메일을 시도합니다.
        log('❌ 예상치 못한 오류가 발생했습니다.')
        traceback.print_exc()
        try:
            send_email(
                f'{MAIL_SUBJECT_PREFIX} ⚠️ 수집 실패 알림',
                f'<pre style="font-size:12px;">{traceback.format_exc()}</pre>',
                None,
            )
        except Exception:
            pass
        sys.exit(1)
